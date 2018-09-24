from math import floor, ceil
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime
import os
# import pandas as pd
# import numpy as np


class TimeTrialUtils:
    template_loader = False
    template_env = False

    def __init__(self, cols):
        self.cols = cols

        self.template_loader = FileSystemLoader(searchpath="templates/")
        self.template_env = Environment(loader=self.template_loader)
        self.template_env.trim_blocks = True
        self.template_env.lstrip_blocks = True

    def create_html(self, filename, delimiter):
        data = []

        # cell headers
        width1 = floor((100 * 4/5) // self.cols)
        width2 = floor((100 * 1/5) // self.cols)
        header = {'width_name': width1, 'width_time': width2}

        f = open(filename, 'r')
        lines = f.readlines()

        num_lines = len(lines)
        diff = int(ceil(num_lines/self.cols))
        # split into columns
        columns = []
        for x in range(0, self.cols):
            start = x * diff
            end = (x+1) * diff
            columns.append(lines[start:end])

        for x in range(0, diff):
            row = []
            for y in range(0, len(columns)):
                if x < len(columns[y]):
                    split = columns[y][x].split(delimiter)
                    name = split[0].strip()
                    result = split[1].replace('\n', '')
                    pb = split[2]
                    desc = split[3]
                    # print(desc)
                    # handle different ways of showing PB or first TT
                    if "**" in desc:
                        result_type = 'first'
                    elif "*" in desc:
                        result_type = 'pb'
                    else:
                        result_type = 'normal'
                    row.append({'name': name, 'time': {'result': result, 'result_type': result_type, 'prev': pb}})
                else:
                    row.append({'name': '', 'time': {'result': '', 'result_type': 'none', 'prev': ''}})

            data.append(row)

        template = self.template_env.get_template("results.html")
        return template.render({'cols': self.cols, 'header': header, 'data': data})

    def export_html(self, results):
        data = []
        # cell headers
        width1 = floor((100 * 4 / 5) // self.cols)
        width2 = floor((100 * 1 / 5) // self.cols)
        header = {'width_name': width1, 'width_time': width2}

        for item in results:
            if item['is_pb'] == 1:
                result_type = 'pb'
            elif item['is_first_time'] == 1:
                result_type = 'first'
            else:
                result_type = 'normal'
            data.append(
                {
                    'name': item['name'],
                    'time': {'result': item['latest'], 'prev': item['pb'], 'result_type': result_type}
                }
            )

        sections = list(TimeTrialUtils.chunks(data, self.cols))

        template = self.template_env.get_template("results.html")
        return template.render({'cols': self.cols, 'header': header, 'data': sections})

    @staticmethod
    def get_sections(seq, num):
        out = []
        last = 0.0

        while last < len(seq):
            out.append(seq[int(last):int(last + num)])
            last += num

        return out

    @staticmethod
    def chunks(seq, chunk_size):
        """Yield successive chunkSize-sized chunks from list."""
        for i in range(0, len(seq), chunk_size):
            yield seq[i:i + chunk_size]

    @staticmethod
    def get_names(name):
        if name == '' or name is None:
            return False

        parts = name.strip().split(' ')
        if len(parts) > 1:
            first_name = parts[0].capitalize()
            last_name = ' '.join(parts[1:])
        else:
            first_name = name
            last_name = ''
        return {'first_name': first_name, 'last_name': last_name}

    @staticmethod
    def get_time(time):
        if type(time) is float or type(time) is int or type(time) is str:
            time = str(time)
            pos = time.find('.')
            time = time.replace('.', ':')
            if pos == -1 and len(time) == 1:
                time = datetime.strptime('0' + time, '%M')
            elif pos == -1 and len(time) == 2:
                time = datetime.strptime(time, '%M')
            elif len(time) == 3:
                time = datetime.strptime('0' + time + '0', '%M:%S')
            elif pos == 2 and len(time) == 4:
                time = datetime.strptime(time + '0', '%M:%S')
            elif len(time) == 4:
                time = datetime.strptime('0' + time, '%M:%S')
            elif len(time) == 5:
                time = datetime.strptime(time, '%M:%S')
            time = time.time()
        return time

    @staticmethod
    def make_active(filename):
        wb = load_workbook(filename)
        ws = wb['Sheet1']
        names = []
        for i in range(2, ws.max_row):
            for j in range(1, ws.max_column):
                cell = get_column_letter(j)+str(i)
                name = ws[cell].value
                if name == '' or name is None:
                    continue
                names.append(TimeTrialUtils.get_names(name))
        return names

    @staticmethod
    def save_as_excel(names, headers, path):
        wb = Workbook()
        ws = wb.active
        row_count = 1
        for col in range(1, len(headers)+1):
            ws[get_column_letter(col) + str(row_count)] = headers[col-1]
        row_count += 1

        for name in names:
            count = 0
            for col in range(1, len(name)+1):
                dict_name = headers[count].lower().replace(' ', '_')
                ws[get_column_letter(col) + str(row_count)] = names[row_count - 2][dict_name]
                count += 1
            row_count += 1
        wb.save(path)


class TimeTrialSpreadsheet:

    wb = False
    runners = {}
    time_trials = {}

    def __init__(self, path):
        if not path:
            abs_path = os.path.abspath(os.path.dirname(__file__))
            # remove /src from path
            path = abs_path[:-4] + '/Run Monash Time Trial.xlsm'
        self.wb = load_workbook(path)

    def get_runners_from(self):
        self.runners = {}
        ws = self.wb['Names']
        data = ws.values
        data = list(data)
        for i in range(10, ws.max_row):
            name = data[i][1]
            if name == '' or name is None:
                continue

            active = data[i][0]
            if active == 'Y':
                active = 1
            else:
                active = 0

            gender = data[i][2]
            if gender == '' or gender is None:
                gender = 'O'
            else:
                gender = gender

            names = TimeTrialUtils.get_names(data[i][1])

            self.runners[i] = {
                'active': active,
                'gender': gender,
                'first_name': names['first_name'].strip(),
                'last_name': names['last_name'].strip(),
            }
        return self.runners

    def get_time_trials_from(self):
        self.time_trials = {}
        # add a place holder for existing PBs
        trial_date = datetime.strptime('01/01/2015', '%d/%m/%Y')
        self.time_trials[3] = {'date': trial_date}

        ws = self.wb['Names']
        data = ws.values
        data = list(data)
        for i in range(1, ws.max_column):
            trial_date = data[7][i]
            if trial_date is None:
                continue

            if type(trial_date) is not datetime:
                trial_date = trial_date.replace("'", "")
                trial_date = datetime.strptime(trial_date, '%d/%m/%Y')

            self.time_trials[i] = {'date': trial_date}

        return self.time_trials

    def get_template_from(self, names, date, path):
        ws = self.wb['Template']

        sheets = ['Names', 'Template2', 'Chart', 'Attendance', 'formulas']
        for sheet in sheets:
            std = self.wb[sheet]
            self.wb.remove(std)

        total_count = len(names)
        cols = 38
        section = 1 + total_count // cols
        pages = 1 + total_count // (cols * 2)

        title = 'TIME TRIAL - '+date.strftime('%A, %B %d, %Y')
        for i in range(1, (pages+1)*6, 6):
            cell = get_column_letter(i)+'1'
            ws[cell] = title

        count = 0
        for j in range(1, section * 3, 3):
            if count >= total_count:
                break
            for i in range(3, 41):
                if count >= total_count:
                    break
                cell = get_column_letter(j)+str(i)
                ws[cell] = names[count][0]
                ws[cell].font = Font(size=12)
                cell = get_column_letter(j+1)+str(i)
                ws[cell] = names[count][1]
                ws[cell].font = Font(size=12)
                count += 1

        self.wb.save(path)

    def get_time_trials_results_from(self):
        time_trial_results = []
        ws = self.wb['Names']
        data = ws.values
        data = list(data)
        for i in range(10, ws.max_row):
            name = data[i][1]
            if name == '' or name is None:
                continue

            runner = self.runners[i]
            for j in range(3, ws.max_column):
                if not data[8][j] == 'TT' and not j == 3:
                    continue
                time = data[i][j]
                if time == '' or time is None:
                    continue

                time = TimeTrialUtils.get_time(time)

                time_trial_date = [v['date'] for k, v in self.time_trials.items() if k == j]

                time_trial_results.append({
                    'time_trial_date': time_trial_date[0],
                    'first_name': runner['first_name'],
                    'last_name': runner['last_name'],
                    'time': time,
                })
        return time_trial_results


class TimeTrialAnalysis:

    def __init__(self):
        print('test')
        # df = pd.DataFrame(data=runners, columns=["active","name",'gender']) # , columns=cols)
        # df = pd.DataFrame(ws.values)
